"""
Bid Parser Evaluation Suite
Usage:
  python tests/eval.py                        # run all documents
  python tests/eval.py --doc "K-26-2431"      # run one document (substring match)
  python tests/eval.py --no-cache             # force re-parse everything
  python tests/eval.py --show-matched         # also print matched rows (verbose)
"""
import argparse
import json
import os
import re
import sys
import requests
import openpyxl
from rapidfuzz import fuzz as _fuzz

# ── Config ────────────────────────────────────────────────────────────────────

BASE_URL   = os.environ.get("PARSER_URL", "http://localhost:8000")
API_KEY    = os.environ.get("PARSER_API_KEY", "")
TESTS_DIR  = os.path.dirname(os.path.abspath(__file__))
GT_DIR     = os.path.join(TESTS_DIR, "ground_truth")
PDF_DIR    = os.path.join(TESTS_DIR, "pdfs")
CACHE_DIR  = os.path.join(TESTS_DIR, "cache")
COL_MAP    = os.path.join(TESTS_DIR, "column_map.json")

FUZZY_THRESHOLD = 0.70

# ── Normalization ─────────────────────────────────────────────────────────────

# All street-type suffixes collapse to "ST" for comparison purposes.
# Whatever Google returns is the real output — but DRIVE=AVENUE=STREET etc. all match in eval.
_SUFFIX_MAP = {
    # Long forms
    "STREET": "ST",   "AVENUE": "ST",   "DRIVE": "ST",   "BOULEVARD": "ST",
    "ROAD": "ST",     "COURT": "ST",    "LANE": "ST",    "PLACE": "ST",
    "WAY": "ST",      "CIRCLE": "ST",   "TERRACE": "ST", "TRAIL": "ST",
    "PARKWAY": "ST",  "HIGHWAY": "ST",  "FREEWAY": "ST",
    # Common abbreviations
    "AVE": "ST",  "AV": "ST",  "DR": "ST",  "BL": "ST",  "BLVD": "ST",
    "RD": "ST",   "CT": "ST",  "LN": "ST",  "PL": "ST",  "WY": "ST",
    "CIR": "ST",  "TER": "ST", "TRL": "ST", "PKWY": "ST", "HWY": "ST",
    "FWY": "ST",  "BI": "ST",  "CR": "ST",
}

# Limit-token synonyms — all mean "end of segment"
_LIMIT_NORM = {
    "EOP": "END", "EOR": "END", "EOS": "END", "EOC": "END",
    "EOL": "END", "EOF": "END", "BEGIN": "END", "BEGINNING": "END",
    "START": "END", "STOP": "END",
    "TEOP": "END",  # OCR artifact: leading char glued to "EOP"
}

# Numeric ordinal → written form (e.g. "5TH" → "FIFTH", "12TH" → "TWELFTH")
_ORDINAL_MAP = {
    "1ST": "FIRST",    "2ND": "SECOND",   "3RD": "THIRD",
    "4TH": "FOURTH",   "5TH": "FIFTH",    "6TH": "SIXTH",
    "7TH": "SEVENTH",  "8TH": "EIGHTH",   "9TH": "NINTH",
    "10TH": "TENTH",   "11TH": "ELEVENTH","12TH": "TWELFTH",
    "13TH": "THIRTEENTH","14TH": "FOURTEENTH","15TH": "FIFTEENTH",
    "16TH": "SIXTEENTH","17TH": "SEVENTEENTH","18TH": "EIGHTEENTH",
    "19TH": "NINETEENTH","20TH": "TWENTIETH",
}

# Cyrillic/Greek lookalikes that DocAI OCR sometimes emits instead of Latin chars
# e.g. "ALMERIA СТ" (Cyrillic С+Т) instead of "ALMERIA CT", "CAM JONAΤΑ" (Greek τ)
_UNICODE_CONFUSABLES = str.maketrans({
    '\u0410': 'A',  # Cyrillic А → A
    '\u0412': 'B',  # Cyrillic В → B
    '\u0421': 'C',  # Cyrillic С → C
    '\u0415': 'E',  # Cyrillic Е → E
    '\u041a': 'K',  # Cyrillic К → K
    '\u041c': 'M',  # Cyrillic М → M
    '\u041d': 'H',  # Cyrillic Н → H
    '\u041e': 'O',  # Cyrillic О → O
    '\u0420': 'R',  # Cyrillic Р → R
    '\u0422': 'T',  # Cyrillic Т → T
    '\u0425': 'X',  # Cyrillic Х → X
    '\u0441': 'c',  # Cyrillic с (lowercase) → c
    '\u03c4': 't',  # Greek τ (tau) → t
    '\u03b1': 'a',  # Greek α (alpha) → a
})

def norm(v) -> str:
    if not v:
        return ""
    s = re.sub(r"\s+", " ", str(v).strip().upper())
    s = s.translate(_UNICODE_CONFUSABLES)  # replace Cyrillic/Greek lookalikes with Latin
    # Strip directional lane suffixes: "ALTON PKWY (EB)" → "ALTON PKWY"
    s = re.sub(r"\s*\((EB|WB|NB|SB)\)\s*$", "", s, flags=re.IGNORECASE)
    # Strip any remaining trailing parenthetical descriptor:
    # "(NORTHBOUND ONLY)", "(BRIDGE OVERPASS)", "(WB ONLY)", "(Ramp N/of Frwy)", etc.
    # Parser correctly drops these; GT sometimes retains them.
    s = re.sub(r"\s*\([^)]+\)\s*$", "", s)
    # Strip asset ID prefixes like "SS-001459-PV1 " before further normalization
    s = re.sub(r'^[A-Z]{1,4}-\d{4,8}-[A-Z0-9]+(?:-[A-Z0-9]+)*\s+', '', s)
    # Normalize all Unicode dash/hyphen variants to ASCII hyphen, then to space
    import unicodedata
    s = "".join("-" if unicodedata.category(c) in ("Pd",) or c in ('\u2010','\u2011','\u2012','\u2013','\u2014','\u2015','\u2212') else c for c in s)
    s = s.replace("-", " ")  # normalize hyphens so "CDS-WEST END" == "CDS - WEST END"
    s = re.sub(r"[^\w\s]", "", s)
    # Normalize doubled limit tokens: "ENDEND" → "END" (OCR artifact)
    s = re.sub(r'\bEND(?:END)+\b', 'END', s, flags=re.IGNORECASE)
    # Split directional prefix glued to ordinal: "S32ND" → "S 32ND", "N15TH" → "N 15TH"
    s = re.sub(r'\b([NSEW])(\d+(?:ST|ND|RD|TH))\b', r'\1 \2', s)
    parts = s.split()
    # Collapse adjacent single-letter tokens: "E H ST" → "EH ST"
    # Handles GT xlsx entries where abbreviated names have extra spaces (e.g. "E   H  ST")
    merged = []
    i = 0
    while i < len(parts):
        if len(parts[i]) == 1 and i + 1 < len(parts) and len(parts[i + 1]) == 1:
            run = parts[i]
            while i + 1 < len(parts) and len(parts[i + 1]) == 1:
                i += 1
                run += parts[i]
            merged.append(run)
        else:
            merged.append(parts[i])
        i += 1
    parts = merged
    # Split tokens like "EJST" → ["EJ", "ST"] to match norm("E J ST") → "EJ ST"
    # Only splits 2-char alpha prefix + known suffix (avoids splitting real words)
    _SPLIT_SUFFIXES = {"ST", "AV", "BL", "RD", "DR", "LN", "CT", "PL", "WY"}
    _NO_SPLIT_2 = {"WE", "IN", "ON", "AT", "BE", "DO", "GO", "NO", "SO", "TO", "UP", "US", "IS"}
    split_parts = []
    for p in parts:
        split = False
        for suf in _SPLIT_SUFFIXES:
            if p.endswith(suf) and len(p) == len(suf) + 2:
                prefix = p[:-len(suf)]
                if prefix.isalpha() and prefix not in _NO_SPLIT_2:
                    split_parts.extend([prefix, suf])
                    split = True
                    break
        if not split:
            split_parts.append(p)
    parts = split_parts

    # Normalize limit tokens (EOP/EOR/EOS/EOF/BEGIN/START → END)
    if len(parts) == 1 and parts[0] in _LIMIT_NORM:
        return _LIMIT_NORM[parts[0]]
    # Normalize ordinals anywhere in the name (e.g. "5TH ST" → "FIFTH ST")
    parts = [_ORDINAL_MAP.get(p, p) for p in parts]
    # Normalize type suffix (last word), with special handling when a compass direction
    # word follows the suffix: "TECHNOLOGY DRIVE WEST" → last="WEST" (not a suffix),
    # but second-to-last "DRIVE" is — normalize it so similarity to "TECHNOLOGY DR" works.
    _COMPASS_WORDS = {"NORTH", "SOUTH", "EAST", "WEST"}
    if parts and parts[-1] in _SUFFIX_MAP:
        parts[-1] = _SUFFIX_MAP[parts[-1]]
    elif len(parts) >= 2 and parts[-1] in _COMPASS_WORDS and parts[-2] in _SUFFIX_MAP:
        parts[-2] = _SUFFIX_MAP[parts[-2]]
    # Strip leading/trailing 1-2 digit noise tokens (row numbers, zone IDs, etc.)
    # Ordinals like "44TH", "21ST" are not pure digits — they stay.
    # Route numbers like 101, 405 are 3+ digits — they stay.
    while parts and parts[0].isdigit() and len(parts[0]) <= 2:
        parts.pop(0)
    while parts and parts[-1].isdigit() and len(parts[-1]) <= 2:
        parts.pop()
    # Normalize CDS variants ("CDS WEST END", "CDS NORTH END", etc.) → "CDS"
    # Hyphen was already collapsed to space above, so "CDS-WEST END" → "CDS WEST END" → "CDS"
    if parts and parts[0] == "CDS":
        return "CDS"
    return " ".join(parts)

_DIR_SUFFIX_RE = re.compile(r"\s*\((EB|WB|NB|SB)\)\s*$", re.IGNORECASE)

def _dir_suffix(s: str) -> str:
    """Extract directional suffix from raw string, e.g. 'ALTON PKWY (WB)' → 'WB'. Empty if none."""
    m = _DIR_SUFFIX_RE.search(str(s or ""))
    return m.group(1).upper() if m else ""

def similarity(a: str, b: str) -> float:
    na, nb = norm(a), norm(b)
    if not na or not nb:
        return 0.0
    return _fuzz.token_sort_ratio(na, nb) / 100.0

def main_street_similarity(a: str, b: str) -> float:
    """Like similarity() but treats conflicting directional suffixes as non-matching."""
    da, db = _dir_suffix(a), _dir_suffix(b)
    # If both sides have a directional suffix and they differ → not the same street
    if da and db and da != db:
        return 0.0
    return similarity(a, b)

def street_key(s: dict) -> tuple:
    return (norm(s.get("main_street", "")),
            norm(s.get("from_street", "")),
            norm(s.get("to_street", "")))

# ── Ground truth loader ───────────────────────────────────────────────────────

def _norm_header(h) -> str:
    """Normalize header for matching: collapse whitespace, uppercase."""
    return re.sub(r"\s+", " ", str(h or "").strip())

def load_ground_truth(doc_key: str, col_map: dict) -> list:
    xlsx_path = os.path.join(GT_DIR, doc_key + ".xlsx")
    if not os.path.exists(xlsx_path):
        print(f"  ⚠  No ground truth xlsx: {xlsx_path}")
        return []

    cfg = col_map.get(doc_key)
    if not cfg:
        print(f"  ⚠  No entry in column_map.json for '{doc_key}' — skipping")
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_rows = []

    for tab_name, tab_cfg_raw in cfg.get("tabs", {}).items():
        # tab_cfg can be a single section dict or a list of section dicts
        sections = tab_cfg_raw if isinstance(tab_cfg_raw, list) else [tab_cfg_raw]

        if tab_name not in wb.sheetnames:
            print(f"  ⚠  Tab '{tab_name}' not found in {doc_key}.xlsx")
            continue

        ws = wb[tab_name]

        for tab_cfg in sections:
            if tab_cfg.get("skip"):
                continue

            header_row_idx = tab_cfg.get("header_row", 2)

            # Build column index map from header row
            col_idx = {}
            for cell in ws[header_row_idx]:
                h = _norm_header(cell.value)
                if h:
                    col_idx[h] = cell.column - 1  # 0-based

            ms_col   = _norm_header(tab_cfg.get("main_street"))
            from_col = _norm_header(tab_cfg.get("from_street")) if tab_cfg.get("from_street") else None
            to_col   = _norm_header(tab_cfg.get("to_street"))   if tab_cfg.get("to_street")   else None
            wt_col   = _norm_header(tab_cfg.get("work_type"))   if tab_cfg.get("work_type")   else None
            lim_col  = _norm_header(tab_cfg.get("limits_col"))  if tab_cfg.get("limits_col")  else None

            if ms_col not in col_idx:
                print(f"  ⚠  main_street col '{ms_col}' not found in tab '{tab_name}' — skipping section")
                continue

            data_start = tab_cfg.get("data_start_row", header_row_idx + 1)
            data_end   = tab_cfg.get("data_end_row", None)
            for row in ws.iter_rows(min_row=data_start, max_row=data_end, values_only=True):
                vals = list(row)

                def get(col_name):
                    if not col_name or col_name not in col_idx:
                        return ""
                    idx = col_idx[col_name]
                    return str(vals[idx]).strip() if idx < len(vals) and vals[idx] is not None else ""

                main = get(ms_col)
                if not main or main in ("None", ""):
                    continue
                # Skip repeated header rows (second table within same tab)
                if norm(main) == norm(ms_col) or norm(main) in ("STREET NAME", "STREET", "NAME"):
                    continue
                # Skip section-header rows like "Fiscal Year 2021-22"
                if main.strip().lower().startswith("fiscal year"):
                    continue

                from_v, to_v = "", ""
                if lim_col:
                    # "FROM to TO" or "FROM TO TO" combined column
                    limits = get(lim_col)
                    parts = re.split(r"\s+to\s+", limits, maxsplit=1, flags=re.IGNORECASE)
                    from_v = parts[0].strip() if len(parts) > 0 else ""
                    to_v   = parts[1].strip() if len(parts) > 1 else ""
                else:
                    from_v = get(from_col) if from_col else ""
                    to_v   = get(to_col)   if to_col   else ""

                wt = get(wt_col) if wt_col else ""

                all_rows.append({
                    "main_street": main,
                    "from_street": from_v,
                    "to_street":   to_v,
                    "work_type":   wt,
                    "_tab": tab_name,
                })

    return all_rows

# ── Parser runner ─────────────────────────────────────────────────────────────

def run_parser(doc_key: str, use_cache: bool = True) -> tuple:
    """Returns (streets, low_confidence_streets)."""
    cache_path = os.path.join(CACHE_DIR, doc_key + ".json")

    if use_cache and os.path.exists(cache_path):
        print(f"  📦 Using cached result")
        with open(cache_path) as f:
            cached = json.load(f)
        # Support old cache format (plain list) and new format (dict with streets key)
        if isinstance(cached, list):
            return cached, []
        return cached.get("streets", []), cached.get("low_confidence_streets", [])

    # PDF has same base name as xlsx
    pdf_path = os.path.join(PDF_DIR, doc_key + ".pdf")
    if not os.path.exists(pdf_path):
        print(f"  ⚠  PDF not found: {pdf_path}")
        return [], []

    print(f"  🔄 Parsing {doc_key}.pdf ...")
    headers = {"X-Api-Key": API_KEY} if API_KEY else {}
    with open(pdf_path, "rb") as f:
        resp = requests.post(
            f"{BASE_URL}/parse",
            headers=headers,
            files={"file": (doc_key + ".pdf", f, "application/pdf")},
            timeout=600,
        )
    resp.raise_for_status()
    body = resp.json()
    streets           = body.get("streets", [])
    low_conf_streets  = body.get("low_confidence_streets", [])

    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump({"streets": streets, "low_confidence_streets": low_conf_streets}, f, indent=2)

    return streets, low_conf_streets

# ── Matching ──────────────────────────────────────────────────────────────────

def match_streets(truth: list, parsed: list) -> dict:
    matched, missed, extra = [], [], []
    used = set()

    # Deduplicate ground truth — use norm() so formatting variants collapse
    # (e.g. "CDS-SOUTH END" == "CDS - SOUTH END" == "CDS SOUTH END" → "CDS").
    seen_keys, deduped_truth = set(), []
    for t in truth:
        k = (
            norm(t.get("main_street") or ""),
            norm(t.get("from_street") or ""),
            norm(t.get("to_street")   or ""),
        )
        if k not in seen_keys:
            seen_keys.add(k)
            deduped_truth.append(t)
    truth = deduped_truth

    # Remove one-sided GT entries (one blank endpoint) when a two-sided entry
    # (both endpoints filled) exists for the same main street.
    # Some GT tabs record only "end location" with "begin location" blank;
    # the complete entry appears in another tab.  Keeping the incomplete version
    # causes it to "steal" a parser match via the swap_ok path, leaving the
    # proper two-sided entry unmatched.
    from collections import defaultdict as _defaultdict
    _by_main = _defaultdict(list)
    for t in truth:
        _by_main[norm(t.get("main_street") or "")].append(t)
    truth = [
        t for t in truth
        if not (
            # one-sided: exactly one of from/to is blank
            bool((t.get("from_street") or "").strip()) != bool((t.get("to_street") or "").strip())
            # AND a two-sided entry exists for the same main street
            and any(
                (g.get("from_street") or "").strip() and (g.get("to_street") or "").strip()
                for g in _by_main[norm(t.get("main_street") or "")]
            )
        )
    ]

    # Deduplicate parsed streets using norm() so CDS variants collapse
    # ("CDS WEST END" == "CDS-WEST END" → same slot).
    # Do NOT collapse swapped orientations (from→to vs to→from) because NB/SB
    # directions are genuinely different segments.
    seen_keys, deduped = set(), []
    for p in parsed:
        k = (
            norm(p.get("main_street") or ""),
            norm(p.get("from_street") or ""),
            norm(p.get("to_street")   or ""),
        )
        if k not in seen_keys:
            seen_keys.add(k)
            deduped.append(p)
    parsed = deduped

    for t in truth:
        tk = street_key(t)
        best_score, best_idx = 0.0, -1
        for i, p in enumerate(parsed):
            if i in used:
                continue
            pk = street_key(p)
            ms_score = main_street_similarity(t.get("main_street", ""), p.get("main_street", ""))
            # Check normal orientation and swapped FROM/TO (limits can be written either direction)
            from_ok  = (not tk[1] or not pk[1] or similarity(tk[1], pk[1]) >= 0.6)
            to_ok    = (not tk[2] or not pk[2] or similarity(tk[2], pk[2]) >= 0.6)
            swap_ok  = ((not tk[1] or not pk[2] or similarity(tk[1], pk[2]) >= 0.6) and
                        (not tk[2] or not pk[1] or similarity(tk[2], pk[1]) >= 0.6))
            if ms_score > best_score and ((from_ok and to_ok) or swap_ok):
                best_score, best_idx = ms_score, i
        if best_score >= FUZZY_THRESHOLD and best_idx >= 0:
            matched.append({"truth": t, "parsed": parsed[best_idx], "score": round(best_score, 2)})
            used.add(best_idx)
        else:
            missed.append(t)

    for i, p in enumerate(parsed):
        if i not in used:
            extra.append(p)

    total = len(truth)
    precision = len(matched) / (len(matched) + len(extra)) if (matched or extra) else 0
    recall    = len(matched) / total if total else 0
    f1        = 2 * precision * recall / (precision + recall) if (precision + recall) else 0

    return {
        "matched": matched, "missed": missed, "extra": extra,
        "precision": precision, "recall": recall, "f1": f1,
        "total_truth": total, "total_parsed": len(parsed),
    }

# ── Report ────────────────────────────────────────────────────────────────────

def print_report(doc_key: str, r: dict, show_matched: bool = False, low_conf: list = None):
    p, re_, f1 = r["precision"], r["recall"], r["f1"]
    bar = "█" * int(f1 * 20) + "░" * (20 - int(f1 * 20))
    print(f"\n{'='*60}")
    print(f"  {doc_key}")
    print(f"{'='*60}")
    lc_note = f"  |  Low-conf (removed): {len(low_conf)}" if low_conf else ""
    print(f"  Truth: {r['total_truth']}  |  Parsed: {r['total_parsed']}  |  "
          f"Matched: {len(r['matched'])}  Missed: {len(r['missed'])}  Extra: {len(r['extra'])}{lc_note}")
    print(f"  Precision: {p:.0%}   Recall: {re_:.0%}   F1: {f1:.0%}  [{bar}]")

    if show_matched and r["matched"]:
        print(f"\n  ✅ MATCHED ({len(r['matched'])}):")
        for m in r["matched"]:
            t, parsed = m["truth"], m["parsed"]
            print(f"     {t['main_street']} | {t.get('from_street','')} → {t.get('to_street','')}  "
                  f"(score {m['score']}, tab: {t.get('_tab','')})")

    if r["missed"]:
        print(f"\n  ❌ MISSED ({len(r['missed'])}):")
        for s in r["missed"][:30]:
            print(f"     {s['main_street']}  |  {s.get('from_street','')} → {s.get('to_street','')}  "
                  f"[{s.get('_tab','')}]")
        if len(r["missed"]) > 30:
            print(f"     ... and {len(r['missed']) - 30} more")

    if r["extra"]:
        print(f"\n  ➕ EXTRA / false positives ({len(r['extra'])}):")
        for s in r["extra"][:15]:
            print(f"     {s.get('main_street','')}  |  {s.get('from_street','')} → {s.get('to_street','')}  "
                  f"[p.{s.get('page','')}]")
        if len(r["extra"]) > 15:
            print(f"     ... and {len(r['extra']) - 15} more")

    if low_conf:
        print(f"\n  🔍 LOW CONFIDENCE (removed from scoring, {len(low_conf)}):")
        for s in low_conf[:20]:
            print(f"     {s.get('main_street','')}  |  {s.get('from_street','')} → {s.get('to_street','')}  "
                  f"[p.{s.get('page','')}]")
        if len(low_conf) > 20:
            print(f"     ... and {len(low_conf) - 20} more")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--doc", help="Run only docs whose key contains this string")
    parser.add_argument("--no-cache", action="store_true", help="Force re-parse all docs")
    parser.add_argument("--show-matched", action="store_true", help="Print matched rows too")
    args = parser.parse_args()

    with open(COL_MAP) as f:
        col_map = json.load(f)

    # All doc keys = xlsx filenames without extension
    all_keys = [
        os.path.splitext(fn)[0]
        for fn in sorted(os.listdir(GT_DIR))
        if fn.endswith(".xlsx")
    ]

    if args.doc:
        all_keys = [k for k in all_keys if args.doc.lower() in k.lower()]
        if not all_keys:
            print(f"No xlsx file matching '{args.doc}' found in tests/ground_truth/")
            sys.exit(1)

    all_results = {}
    for doc_key in all_keys:
        print(f"\n▶ {doc_key}")
        truth  = load_ground_truth(doc_key, col_map)
        if not truth:
            print(f"  ⚠  No ground truth rows loaded — skipping")
            continue
        print(f"  📋 {len(truth)} ground truth rows loaded")
        parsed, low_conf = run_parser(doc_key, use_cache=not args.no_cache)
        result = match_streets(truth, parsed)
        all_results[doc_key] = result
        print_report(doc_key, result, show_matched=args.show_matched, low_conf=low_conf)

    if len(all_results) > 1:
        total_matched = sum(len(r["matched"]) for r in all_results.values())
        total_missed  = sum(len(r["missed"])  for r in all_results.values())
        total_extra   = sum(len(r["extra"])   for r in all_results.values())
        total_truth   = sum(r["total_truth"]  for r in all_results.values())
        total_parsed  = sum(r["total_parsed"] for r in all_results.values())
        op = total_matched / (total_matched + total_extra) if (total_matched + total_extra) else 0
        or_ = total_matched / total_truth if total_truth else 0
        of1 = 2 * op * or_ / (op + or_) if (op + or_) else 0
        bar = "█" * int(of1 * 20) + "░" * (20 - int(of1 * 20))
        print(f"\n{'='*60}")
        print(f"  OVERALL  ({len(all_results)} documents)")
        print(f"{'='*60}")
        print(f"  Truth: {total_truth}  |  Parsed: {total_parsed}  |  "
              f"Matched: {total_matched}  Missed: {total_missed}  Extra: {total_extra}")
        print(f"  Precision: {op:.0%}   Recall: {or_:.0%}   F1: {of1:.0%}  [{bar}]")

        # Per-doc summary table
        print(f"\n  {'Document':<55} {'P':>5} {'R':>5} {'F1':>5}")
        print(f"  {'-'*70}")
        for dk, r in sorted(all_results.items(), key=lambda x: x[1]['f1']):
            print(f"  {dk:<55} {r['precision']:>4.0%} {r['recall']:>4.0%} {r['f1']:>4.0%}")

if __name__ == "__main__":
    main()
